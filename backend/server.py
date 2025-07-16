from fastapi import FastAPI, APIRouter, HTTPException, File, UploadFile, Form
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import uuid
from datetime import datetime
import json
import asyncio
import aiofiles
import io
from docx import Document
import PyPDF2
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# AI imports
import openai
import anthropic
import google.generativeai as genai

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

# Pydantic Models
class AIProviderConfig(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    provider: str  # 'openai', 'anthropic', 'google'
    api_key: str
    model: str
    max_tokens: int = 4000
    temperature: float = 0.7
    created_at: datetime = Field(default_factory=datetime.utcnow)
    is_active: bool = True

class AIProviderConfigCreate(BaseModel):
    provider: str
    api_key: str
    model: str
    max_tokens: int = 4000
    temperature: float = 0.7

class TestCase(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    description: str
    preconditions: str
    steps: List[str]
    expected_result: str
    priority: str = "Medium"  # Low, Medium, High
    category: str = "Functional"
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)
    is_selected: bool = False

class TestCaseCreate(BaseModel):
    title: str
    description: str
    preconditions: str
    steps: List[str]
    expected_result: str
    priority: str = "Medium"
    category: str = "Functional"

class TestCaseUpdate(BaseModel):
    title: Optional[str] = None
    description: Optional[str] = None
    preconditions: Optional[str] = None
    steps: Optional[List[str]] = None
    expected_result: Optional[str] = None
    priority: Optional[str] = None
    category: Optional[str] = None
    is_selected: Optional[bool] = None

class TestCaseGenerationRequest(BaseModel):
    prompt: str
    context_files: Optional[List[str]] = []
    requirements: Optional[str] = ""
    test_type: str = "Functional"
    num_test_cases: int = 5

class Project(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    name: str
    description: str
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class ProjectCreate(BaseModel):
    name: str
    description: str

class Transcript(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    title: str
    content: str
    meeting_date: Optional[str] = None
    participants: Optional[str] = None
    created_at: datetime = Field(default_factory=datetime.utcnow)
    updated_at: datetime = Field(default_factory=datetime.utcnow)

class TranscriptCreate(BaseModel):
    title: str
    content: str
    meeting_date: Optional[str] = None
    participants: Optional[str] = None

# AI Provider Management
class AIProviderManager:
    def __init__(self):
        self.providers = {}
    
    async def get_active_provider(self) -> Optional[AIProviderConfig]:
        """Get the active AI provider configuration"""
        config = await db.ai_configs.find_one({"is_active": True})
        if config:
            return AIProviderConfig(**config)
        return None
    
    async def generate_test_cases(self, request: TestCaseGenerationRequest, file_contents: List[str] = None) -> List[TestCase]:
        """Generate test cases using the active AI provider"""
        provider_config = await self.get_active_provider()
        if not provider_config:
            raise HTTPException(status_code=400, detail="No active AI provider configured")
        
        # Build context from files
        context = ""
        if file_contents:
            context = "\n\n".join([f"File content:\n{content}" for content in file_contents])
        
        # Add transcript context if provided
        transcript_context = ""
        if hasattr(request, 'selected_transcripts') and request.selected_transcripts:
            transcript_docs = await db.transcripts.find({"id": {"$in": request.selected_transcripts}}).to_list(100)
            if transcript_docs:
                transcript_context = "\n\n".join([
                    f"Meeting Transcript - {doc['title']}:\n{doc['content']}"
                    for doc in transcript_docs
                ])
        
        # Create the prompt
        system_prompt = f"""You are an expert QA engineer specialized in creating comprehensive test cases. 
        
Generate {request.num_test_cases} detailed test cases based on the following requirements:

Requirements: {request.prompt}
Test Type: {request.test_type}

Context from uploaded files:
{context}

Meeting transcripts context:
{transcript_context}

For each test case, provide:
1. Title - Clear, descriptive title
2. Description - Brief description of what is being tested
3. Preconditions - What needs to be set up before testing
4. Steps - Detailed step-by-step instructions (as array)
5. Expected Result - What should happen if the test passes
6. Priority - High, Medium, or Low
7. Category - Functional, Performance, Security, Usability, etc.

Return the response as a JSON array of test cases with the exact structure:
[
  {{
    "title": "Test case title",
    "description": "Test case description",
    "preconditions": "Prerequisites for the test",
    "steps": ["Step 1", "Step 2", "Step 3"],
    "expected_result": "Expected outcome",
    "priority": "Medium",
    "category": "Functional"
  }}
]

Make sure the test cases are realistic, actionable, and cover different scenarios including positive, negative, and edge cases."""

        try:
            if provider_config.provider == 'openai':
                openai.api_key = provider_config.api_key
                response = openai.chat.completions.create(
                    model=provider_config.model,
                    messages=[
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": f"Generate {request.num_test_cases} test cases for: {request.prompt}"}
                    ],
                    max_tokens=provider_config.max_tokens,
                    temperature=provider_config.temperature
                )
                content = response.choices[0].message.content
                
            elif provider_config.provider == 'anthropic':
                client = anthropic.Anthropic(api_key=provider_config.api_key)
                response = client.messages.create(
                    model=provider_config.model,
                    max_tokens=provider_config.max_tokens,
                    temperature=provider_config.temperature,
                    system=system_prompt,
                    messages=[
                        {"role": "user", "content": f"Generate {request.num_test_cases} test cases for: {request.prompt}"}
                    ]
                )
                content = response.content[0].text
                
            elif provider_config.provider == 'google':
                genai.configure(api_key=provider_config.api_key)
                model = genai.GenerativeModel(provider_config.model)
                response = model.generate_content(
                    f"{system_prompt}\n\nGenerate {request.num_test_cases} test cases for: {request.prompt}",
                    generation_config=genai.types.GenerationConfig(
                        max_output_tokens=provider_config.max_tokens,
                        temperature=provider_config.temperature,
                    )
                )
                content = response.text
            
            # Parse the JSON response
            try:
                # Clean the response to extract JSON
                content = content.strip()
                if content.startswith('```json'):
                    content = content[7:-3]
                elif content.startswith('```'):
                    content = content[3:-3]
                
                test_cases_data = json.loads(content)
                
                # Convert to TestCase objects
                test_cases = []
                for tc_data in test_cases_data:
                    test_case = TestCase(**tc_data)
                    test_cases.append(test_case)
                
                return test_cases
                
            except json.JSONDecodeError as e:
                logger.error(f"Failed to parse AI response as JSON: {e}")
                logger.error(f"Response content: {content}")
                raise HTTPException(status_code=500, detail="Failed to parse AI response")
                
        except Exception as e:
            logger.error(f"AI generation failed: {e}")
            raise HTTPException(status_code=500, detail=f"AI generation failed: {str(e)}")

ai_manager = AIProviderManager()

# File processing utilities
async def process_uploaded_file(file: UploadFile) -> str:
    """Process uploaded file and extract text content"""
    try:
        content = await file.read()
        
        if file.filename.endswith('.txt'):
            return content.decode('utf-8')
        
        elif file.filename.endswith('.pdf'):
            pdf_reader = PyPDF2.PdfReader(io.BytesIO(content))
            text = ""
            for page in pdf_reader.pages:
                text += page.extract_text()
            return text
        
        elif file.filename.endswith('.docx'):
            doc = Document(io.BytesIO(content))
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text
        
        else:
            raise HTTPException(status_code=400, detail="Unsupported file format")
            
    except Exception as e:
        logger.error(f"File processing failed: {e}")
        raise HTTPException(status_code=500, detail=f"File processing failed: {str(e)}")

# API Routes

# AI Provider Configuration
@api_router.post("/ai-providers", response_model=AIProviderConfig)
async def create_ai_provider(config: AIProviderConfigCreate):
    """Create or update AI provider configuration"""
    # Deactivate existing providers
    await db.ai_configs.update_many({}, {"$set": {"is_active": False}})
    
    # Create new active provider
    provider_dict = config.dict()
    provider_obj = AIProviderConfig(**provider_dict)
    await db.ai_configs.insert_one(provider_obj.dict())
    return provider_obj

@api_router.get("/ai-providers", response_model=List[AIProviderConfig])
async def get_ai_providers():
    """Get all AI provider configurations"""
    providers = await db.ai_configs.find().to_list(1000)
    return [AIProviderConfig(**provider) for provider in providers]

@api_router.get("/ai-providers/active", response_model=Optional[AIProviderConfig])
async def get_active_ai_provider():
    """Get active AI provider configuration"""
    return await ai_manager.get_active_provider()

# Test Case Generation
@api_router.post("/generate-test-cases", response_model=List[TestCase])
async def generate_test_cases(
    prompt: str = Form(...),
    test_type: str = Form("Functional"),
    num_test_cases: int = Form(5),
    selected_transcripts: str = Form("[]"),
    selected_alm: str = Form(""),
    selected_alm_items: str = Form("[]"),
    files: List[UploadFile] = File(default=[])
):
    """Generate test cases using AI"""
    
    # Process uploaded files
    file_contents = []
    for file in files:
        if file.filename:
            content = await process_uploaded_file(file)
            file_contents.append(content)
    
    # Parse selected transcripts
    try:
        transcript_ids = json.loads(selected_transcripts)
    except:
        transcript_ids = []
    
    # Create generation request
    request = TestCaseGenerationRequest(
        prompt=prompt,
        test_type=test_type,
        num_test_cases=num_test_cases
    )
    
    # Add transcript context to request for processing
    request.selected_transcripts = transcript_ids
    
    # Generate test cases
    test_cases = await ai_manager.generate_test_cases(request, file_contents)
    
    # Save to database
    for test_case in test_cases:
        await db.test_cases.insert_one(test_case.dict())
    
    return test_cases

# Test Case Management
@api_router.get("/test-cases", response_model=List[TestCase])
async def get_test_cases():
    """Get all test cases"""
    test_cases = await db.test_cases.find().sort("created_at", -1).to_list(1000)
    return [TestCase(**tc) for tc in test_cases]

@api_router.get("/test-cases/{test_case_id}", response_model=TestCase)
async def get_test_case(test_case_id: str):
    """Get specific test case"""
    test_case = await db.test_cases.find_one({"id": test_case_id})
    if not test_case:
        raise HTTPException(status_code=404, detail="Test case not found")
    return TestCase(**test_case)

@api_router.put("/test-cases/{test_case_id}", response_model=TestCase)
async def update_test_case(test_case_id: str, update_data: TestCaseUpdate):
    """Update test case"""
    update_dict = {k: v for k, v in update_data.dict().items() if v is not None}
    update_dict["updated_at"] = datetime.utcnow()
    
    result = await db.test_cases.update_one(
        {"id": test_case_id},
        {"$set": update_dict}
    )
    
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Test case not found")
    
    updated_test_case = await db.test_cases.find_one({"id": test_case_id})
    return TestCase(**updated_test_case)

@api_router.delete("/test-cases/{test_case_id}")
async def delete_test_case(test_case_id: str):
    """Delete test case"""
    result = await db.test_cases.delete_one({"id": test_case_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Test case not found")
    return {"message": "Test case deleted successfully"}

@api_router.post("/test-cases/bulk-select")
async def bulk_select_test_cases(test_case_ids: List[str]):
    """Bulk select/deselect test cases"""
    await db.test_cases.update_many(
        {"id": {"$in": test_case_ids}},
        {"$set": {"is_selected": True}}
    )
    return {"message": f"Selected {len(test_case_ids)} test cases"}

@api_router.delete("/test-cases")
async def delete_all_test_cases():
    """Delete all test cases"""
    result = await db.test_cases.delete_many({})
    return {"message": f"Deleted {result.deleted_count} test cases"}

# Export functionality
@api_router.get("/export/excel")
async def export_to_excel():
    """Export selected test cases to Excel"""
    test_cases = await db.test_cases.find({"is_selected": True}).to_list(1000)
    
    if not test_cases:
        raise HTTPException(status_code=400, detail="No test cases selected for export")
    
    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"
    
    # Headers
    headers = ["ID", "Title", "Description", "Preconditions", "Steps", "Expected Result", "Priority", "Category", "Created At"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Data rows
    for row, tc in enumerate(test_cases, 2):
        ws.cell(row=row, column=1, value=tc["id"])
        ws.cell(row=row, column=2, value=tc["title"])
        ws.cell(row=row, column=3, value=tc["description"])
        ws.cell(row=row, column=4, value=tc["preconditions"])
        ws.cell(row=row, column=5, value="\n".join(tc["steps"]))
        ws.cell(row=row, column=6, value=tc["expected_result"])
        ws.cell(row=row, column=7, value=tc["priority"])
        ws.cell(row=row, column=8, value=tc["category"])
        ws.cell(row=row, column=9, value=tc["created_at"].strftime("%Y-%m-%d %H:%M:%S"))
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        io.BytesIO(output.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=test_cases.xlsx"}
    )

@api_router.get("/export/json")
async def export_to_json():
    """Export selected test cases to JSON"""
    test_cases = await db.test_cases.find({"is_selected": True}).to_list(1000)
    
    if not test_cases:
        raise HTTPException(status_code=400, detail="No test cases selected for export")
    
    # Convert to JSON serializable format
    export_data = []
    for tc in test_cases:
        tc_dict = dict(tc)
        tc_dict["created_at"] = tc_dict["created_at"].isoformat()
        tc_dict["updated_at"] = tc_dict["updated_at"].isoformat()
        export_data.append(tc_dict)
    
    json_data = json.dumps(export_data, indent=2)
    
    return StreamingResponse(
        io.BytesIO(json_data.encode()),
        media_type="application/json",
        headers={"Content-Disposition": "attachment; filename=test_cases.json"}
    )

# Transcript Management
@api_router.post("/transcripts", response_model=Transcript)
async def create_transcript(transcript: TranscriptCreate):
    """Create a new transcript"""
    transcript_dict = transcript.dict()
    transcript_obj = Transcript(**transcript_dict)
    await db.transcripts.insert_one(transcript_obj.dict())
    return transcript_obj

@api_router.get("/transcripts", response_model=List[Transcript])
async def get_transcripts():
    """Get all transcripts"""
    transcripts = await db.transcripts.find().sort("created_at", -1).to_list(1000)
    return [Transcript(**t) for t in transcripts]

@api_router.get("/transcripts/{transcript_id}", response_model=Transcript)
async def get_transcript(transcript_id: str):
    """Get specific transcript"""
    transcript = await db.transcripts.find_one({"id": transcript_id})
    if not transcript:
        raise HTTPException(status_code=404, detail="Transcript not found")
    return Transcript(**transcript)

@api_router.delete("/transcripts/{transcript_id}")
async def delete_transcript(transcript_id: str):
    """Delete transcript"""
    result = await db.transcripts.delete_one({"id": transcript_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Transcript not found")
    return {"message": "Transcript deleted successfully"}

@api_router.post("/transcripts/upload")
async def upload_transcripts(files: List[UploadFile] = File(...)):
    """Upload transcript files"""
    transcripts = []
    
    for file in files:
        if file.filename and file.filename.endswith('.txt'):
            content = await file.read()
            text_content = content.decode('utf-8')
            
            # Create transcript from file
            transcript = Transcript(
                title=file.filename.replace('.txt', ''),
                content=text_content
            )
            
            await db.transcripts.insert_one(transcript.dict())
            transcripts.append(transcript)
    
    return {"message": f"Uploaded {len(transcripts)} transcripts", "transcripts": transcripts}

# Health check
@api_router.get("/")
async def root():
    return {"message": "Gen Studio AI API is running"}

@api_router.get("/health")
async def health_check():
    return {"status": "healthy", "timestamp": datetime.utcnow().isoformat()}

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()